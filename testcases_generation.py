from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def create_excel():
    # Test cases data
    test_cases = [
        ["TC001", "Insert Card and Enter PIN", "High", "1. Insert card\n2. Enter PIN", "", "Successful login", "", "2024-03-14", "John Doe"],
        ["TC002", "Withdraw Cash", "Medium", "1. Select 'Withdraw Cash'\n2. Enter amount\n3. Take cash", "", "Cash dispensed", "", "2024-03-14", "Jane Smith"],
        # Add more test cases as needed
    ]

    # Create workbook and sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ATM Test Cases"

    # Header style
    header_font = Font(color="FFFFFF")
    header_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
    for cell in sheet["A1:I1"]:
        cell.font = header_font
        cell.fill = header_fill

    # Header cells
    headers = ["Test Case Id", "Test Case Description", "Priority", "Test Steps", "ALM path to Upload", "Expected Result", "Actual Result", "Created On", "Created By"]
    for i, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=i, value=header)

    # Add test cases
    for test_case in test_cases:
        sheet.append(test_case)

    # Auto-size columns
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    # Save workbook
    workbook.save("ATM_Test_Cases.xlsx")

if __name__ == "__main__":
    create_excel()
