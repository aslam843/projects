import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import os

def merge_excel_files(file1, file2, key_column, output_file):

    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    common_columns = list(set(df1.columns) & set(df2.columns))

    common_columns.remove(key_column)
    merged_df = pd.DataFrame({key_column: df1[key_column]})

    for column in common_columns:
        merged_df[column] = df1[column]
        merged_df[f'{column}_'] = df2[column]
    merged_df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    for column in common_columns:
        for row in range(2, ws.max_row + 1):  # Start from row 2
            cell1 = ws.cell(row=row, column=merged_df.columns.get_loc(column) + 1)
            cell2 = ws.cell(row=row, column=merged_df.columns.get_loc(f'{column}_') + 1)
            if cell1.value is not None and cell2.value is not None and cell1.value != cell2.value:
                font = Font(color="FF0000")  # Red font color
                cell1.font = font
    wb.save(output_file)
    print(f'Merged file saved as {output_file}')

if __name__ == "__main__":
    file1 = 'DIY.xlsx'
    file2 = 'Seller.xlsx'
    key_column = 'ID'
    output_file = 'merged_output.xlsx'

    merge_excel_files(file1, file2, key_column, output_file)
